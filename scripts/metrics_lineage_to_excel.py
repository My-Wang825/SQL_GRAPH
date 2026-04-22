# -*- coding: utf-8 -*-
"""
将指标平台 queryAll 返回的 JSON 落盘，并导出为多 Sheet 的 Excel，便于对照 vertexType 含义。

用法（项目根目录执行）::

    # 方式 A：从环境变量拉取（与 backend/config 一致），写出 JSON + xlsx
    python scripts/metrics_lineage_to_excel.py --fetch

    # 方式 B：已有原始 JSON 文件，只生成 Excel（可同时再写一份规范化 raw 副本）
    python scripts/metrics_lineage_to_excel.py --json path/to/lineage.json

环境变量（--fetch 时）：METRICS_LINEAGE_QUERY_ALL_URL、METRICS_TENANT_ID、
METRICS_AUTH_TYPE、METRICS_AUTH_VALUE、METRICS_HTTP_TIMEOUT_SEC

输出目录默认：项目根下 exports/metrics_lineage/
  - lineage_queryAll_raw.json   平台原始响应
  - lineage_vertex_edges.xlsx   顶点 / 边 / 类型汇总
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, DefaultDict, Dict, List, Optional, Set, Tuple

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def _load_payload_from_json(path: Path) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, dict):
        raise SystemExit("JSON 根节点必须是对象")
    return data


def _iter_edges(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    data = payload.get("data") or {}
    edge_list = data.get("edgeList") or data.get("edge_list") or []
    return edge_list if isinstance(edge_list, list) else []


def _collect_vertices_and_edges(
    edge_list: List[Dict[str, Any]],
) -> Tuple[
    Dict[Tuple[str, str], Dict[str, Any]],
    List[Dict[str, Any]],
    DefaultDict[Tuple[str, str], int],
    DefaultDict[Tuple[str, str], int],
]:
    """唯一顶点 (vertex_id, vertex_type) -> 记录；边列表；作为 src/dst 出现次数。"""
    vertices: Dict[Tuple[str, str], Dict[str, Any]] = {}
    src_cnt: DefaultDict[Tuple[str, str], int] = defaultdict(int)
    dst_cnt: DefaultDict[Tuple[str, str], int] = defaultdict(int)
    edges_out: List[Dict[str, Any]] = []

    def norm_v(raw: Any) -> Optional[Tuple[str, str, str]]:
        if not isinstance(raw, dict):
            return None
        vid = (raw.get("vertexId") or raw.get("vertex_id") or "").strip()
        vtype = (raw.get("vertexType") or raw.get("vertex_type") or "").strip()
        tid = (raw.get("tenantId") or raw.get("tenant_id") or "").strip()
        if not vid or not vtype:
            return None
        return vid, vtype, tid

    for edge in edge_list:
        if not isinstance(edge, dict):
            continue
        sv = edge.get("srcVertex") or edge.get("src_vertex") or {}
        dv = edge.get("dstVertex") or edge.get("dst_vertex") or {}
        ns = norm_v(sv)
        nd = norm_v(dv)
        if not ns or not nd:
            continue
        s_id, s_type, s_tid = ns
        t_id, t_type, t_tid = nd
        sk = (s_id, s_type)
        tk = (t_id, t_type)
        if sk not in vertices:
            vertices[sk] = {
                "vertex_id": s_id,
                "vertex_type": s_type,
                "tenant_id": s_tid,
                "id_dot_count": s_id.count("."),
                "id_last_segment": s_id.rsplit(".", 1)[-1] if s_id else "",
            }
        if tk not in vertices:
            vertices[tk] = {
                "vertex_id": t_id,
                "vertex_type": t_type,
                "tenant_id": t_tid,
                "id_dot_count": t_id.count("."),
                "id_last_segment": t_id.rsplit(".", 1)[-1] if t_id else "",
            }
        src_cnt[sk] += 1
        dst_cnt[tk] += 1
        edges_out.append(
            {
                "src_vertex_id": s_id,
                "src_vertex_type": s_type,
                "src_tenant_id": s_tid,
                "dst_vertex_id": t_id,
                "dst_vertex_type": t_type,
                "dst_tenant_id": t_tid,
                "edge_type": edge.get("edgeType") or edge.get("edge_type") or "",
            }
        )

    return vertices, edges_out, src_cnt, dst_cnt


def _type_summary(
    vertices: Dict[Tuple[str, str], Dict[str, Any]],
) -> List[Tuple[str, int, str]]:
    by_type: DefaultDict[str, List[str]] = defaultdict(list)
    for (_vid, vtype), rec in vertices.items():
        by_type[vtype].append(rec["vertex_id"])
    rows: List[Tuple[str, int, str]] = []
    for vtype in sorted(by_type.keys(), key=lambda t: (-len(by_type[t]), t)):
        ids = sorted(by_type[vtype], key=str.lower)
        sample = ids[:8]
        rows.append((vtype, len(ids), " | ".join(sample)))
    return rows


def _autosize_columns(ws, max_width: int = 80) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 12
        for cell in col:
            if cell.value is None:
                continue
            width = min(max_width, max(width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def build_workbook(
    payload: Dict[str, Any],
    vertices: Dict[Tuple[str, str], Dict[str, Any]],
    edges: List[Dict[str, Any]],
    src_cnt: DefaultDict[Tuple[str, str], int],
    dst_cnt: DefaultDict[Tuple[str, str], int],
) -> Workbook:
    wb = Workbook()
    # --- 说明 ---
    ws0 = wb.active
    ws0.title = "使用说明"
    lines = [
        "本文件由 scripts/metrics_lineage_to_excel.py 从 queryAll 原始 JSON 生成。",
        "",
        "工作表说明：",
        "  「类型汇总」：各 vertexType 的去重顶点数量 + 若干 vertexId 示例（用 | 分隔）。",
        "  「顶点清单」：边端点去重后的全量顶点；appear_as_src / appear_as_dst 为在 edgeList 中出现的次数。",
        "  「边清单」：与平台 data.edgeList 一一对应（便于筛选 src/dst 类型组合）。",
        "  「平台元信息」：success、code、traceId、errorMsg 等顶层字段。",
        "",
        "判断 PHYSICAL_TABLE / DATASET 等含义时，建议结合「顶点清单」的 id_last_segment、",
        "以及「边清单」中与您业务相关的若干行一起对照。",
    ]
    for i, text in enumerate(lines, start=1):
        c = ws0.cell(row=i, column=1, value=text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws0.column_dimensions["A"].width = 110

    # --- 类型汇总 ---
    ws1 = wb.create_sheet("类型汇总", 1)
    ws1.append(["vertex_type", "unique_vertex_count", "sample_vertex_ids"])
    for row in _type_summary(vertices):
        ws1.append(list(row))
    header_font = Font(bold=True)
    for cell in ws1[1]:
        cell.font = header_font
    _autosize_columns(ws1)

    # --- 顶点清单 ---
    ws2 = wb.create_sheet("顶点清单", 2)
    headers = [
        "vertex_id",
        "vertex_type",
        "tenant_id",
        "id_dot_count",
        "id_last_segment",
        "appear_as_src",
        "appear_as_dst",
    ]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = header_font
    sorted_keys = sorted(vertices.keys(), key=lambda k: (k[1].lower(), k[0].lower()))
    for vid, vtype in sorted_keys:
        rec = vertices[(vid, vtype)]
        key = (vid, vtype)
        ws2.append(
            [
                rec["vertex_id"],
                rec["vertex_type"],
                rec["tenant_id"],
                rec["id_dot_count"],
                rec["id_last_segment"],
                src_cnt[key],
                dst_cnt[key],
            ]
        )
    _autosize_columns(ws2)

    # --- 边清单 ---
    ws3 = wb.create_sheet("边清单", 3)
    eh = [
        "src_vertex_id",
        "src_vertex_type",
        "src_tenant_id",
        "dst_vertex_id",
        "dst_vertex_type",
        "dst_tenant_id",
        "edge_type",
    ]
    ws3.append(eh)
    for cell in ws3[1]:
        cell.font = header_font
    for e in edges:
        ws3.append(
            [
                e["src_vertex_id"],
                e["src_vertex_type"],
                e["src_tenant_id"],
                e["dst_vertex_id"],
                e["dst_vertex_type"],
                e["dst_tenant_id"],
                e["edge_type"],
            ]
        )
    _autosize_columns(ws3)

    # --- 平台元信息 ---
    ws4 = wb.create_sheet("平台元信息", 4)
    ws4.append(["field", "value"])
    for cell in ws4[1]:
        cell.font = header_font
    meta_keys = [
        "success",
        "code",
        "errorMsg",
        "detailErrorMsg",
        "traceId",
    ]
    for k in meta_keys:
        ws4.append([k, json.dumps(payload.get(k), ensure_ascii=False)])
    ws4.append(["data.edgeList_length", len(edges)])
    _autosize_columns(ws4)

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description="queryAll JSON 落盘并导出 Excel")
    parser.add_argument(
        "--json",
        type=str,
        default="",
        help="已有平台原始 JSON 文件路径（与 --fetch 二选一）",
    )
    parser.add_argument(
        "--fetch",
        action="store_true",
        help="按环境变量从指标平台拉取 queryAll",
    )
    parser.add_argument(
        "--out-dir",
        type=str,
        default="",
        help=f"输出目录，默认 {ROOT / 'exports' / 'metrics_lineage'}",
    )
    args = parser.parse_args()

    if bool(args.json) == bool(args.fetch):
        parser.error("请指定其一：--json <文件> 或 --fetch")

    out_dir = Path(args.out_dir.strip()) if (args.out_dir or "").strip() else ROOT / "exports" / "metrics_lineage"
    out_dir.mkdir(parents=True, exist_ok=True)

    if args.fetch:
        os.chdir(ROOT)
        from backend.metrics.client import MetricsLineageClientError, fetch_query_all_lineage
        from backend.config import (
            METRICS_AUTH_TYPE,
            METRICS_AUTH_VALUE,
            METRICS_HTTP_TIMEOUT_SEC,
            METRICS_LINEAGE_QUERY_ALL_URL,
            METRICS_TENANT_ID,
        )
        if not METRICS_LINEAGE_QUERY_ALL_URL or not METRICS_TENANT_ID or not METRICS_AUTH_VALUE:
            raise SystemExit(
                "缺少环境变量：METRICS_LINEAGE_QUERY_ALL_URL、METRICS_TENANT_ID、METRICS_AUTH_VALUE"
            )
        try:
            payload = fetch_query_all_lineage(
                METRICS_LINEAGE_QUERY_ALL_URL,
                METRICS_TENANT_ID,
                METRICS_AUTH_TYPE,
                METRICS_AUTH_VALUE,
                timeout_sec=METRICS_HTTP_TIMEOUT_SEC,
            )
        except MetricsLineageClientError as e:
            raise SystemExit(str(e)) from e
    else:
        json_path = Path(args.json)
        if not json_path.is_file():
            raise SystemExit(f"文件不存在: {json_path}")
        payload = _load_payload_from_json(json_path)

    raw_path = out_dir / "lineage_queryAll_raw.json"
    with open(raw_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"已写入原始 JSON: {raw_path}")

    edge_list = _iter_edges(payload)
    vertices, edges, src_cnt, dst_cnt = _collect_vertices_and_edges(edge_list)
    wb = build_workbook(payload, vertices, edges, src_cnt, dst_cnt)

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    xlsx_path = out_dir / f"lineage_vertex_edges_{ts}.xlsx"
    wb.save(xlsx_path)
    print(f"已写入 Excel: {xlsx_path}")
    print(f"顶点数(去重): {len(vertices)}  边数: {len(edges)}")


if __name__ == "__main__":
    main()
